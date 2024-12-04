import openpyxl
import sys
from basedef import get_lesson, get_day, get_lesson_num
from dbfunc import get_cab_db
from openpyxl.styles import PatternFill

days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
num = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
num_time = {
    1:"9:00-9:45",
    2:"9:55-10:40",
    3:"11:00-11:45",
    4:"12:05-12:50",
    5:"13:10-13:55",
    6:"14:15-15:00",
    7:"15:10-15:55",
    8:"16:05-16:50",
    9:"17:00-17:45",
    10:"17:55-18:40",
}
class ScheObj():
    row_id = 0
    col_id = 0
    is_tec = False

def find_cabinet(filePath, row_id, col_id):
    wb = openpyxl.load_workbook(filePath, read_only=False, keep_vba=True)

    ws = wb.active
    cv = str(ws.cell(row=int(row_id), column=int(col_id)).value).strip()
    teacher = str(cv.split("/")[-1]).strip()
    teacher = teacher.split(" ")[0]
    print(teacher)
    iter = 0
    for sheet_name in wb.sheetnames:
        
        ws = wb[sheet_name]
        she_arr = []
        lesson = get_lesson(ws)
        for less in lesson:
            if teacher in less.teacher:
                scheObj = ScheObj()
                scheObj.row_id = less.lesson_num + 1
                scheObj.col_id = days.index(less.lesson_day) + 2
                scheObj.is_tec = True
                print(scheObj.row_id)
                print(scheObj.col_id)
                print(scheObj.is_tec)
                she_arr.append(scheObj)

        if teacher not in wb.sheetnames:
            ws = wb.create_sheet(teacher)
        ws = wb[teacher]
        ws.cell(1+iter, 1).value = str(sheet_name)
        for i in range(2, len(days)+2):
            ws.cell(1+iter, i).value = days[i-2]
        for i in range(2, 12):
            ws.cell(i+iter, 1).value = i-1
        

        for obj in she_arr:
            ws.cell(obj.row_id+iter, obj.col_id).value = num_time[obj.row_id-1]
            ws.cell(obj.row_id+iter, obj.col_id).fill = PatternFill(patternType='solid', fgColor='FF3300')
        for rows in ws.iter_rows(max_row=11+iter, max_col=7, min_row=iter):
            for cell in rows:
                if cell.value is None:
                    cell.value = "Свободен"
                    cell.fill = PatternFill(patternType='solid', fgColor='00FF00')
        iter += 11
                  

   
    
    wb.save(filePath)
    print("Успех!")


def main():
    if sys.argv is not None and len(sys.argv) > 1:
        filePath = sys.argv[1]
        find_cabinet(filePath, sys.argv[2], sys.argv[3])

if __name__ == '__main__':
    main()
x = input()